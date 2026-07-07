from __future__ import annotations
import copy
import hashlib
import json
import math
import os
import subprocess
import sys
import time
from collections import Counter
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook, Workbook

os.environ.setdefault('LARK_CLI_TIMEOUT','1800')
repo=Path('/Users/lilixiaoran/工作/转转/ZZ-AI-Business-Analysis-base-migration')
sys.path.insert(0, str(repo))
base_migration=__import__('skills.workflows.机型周数据.base_migration', fromlist=[
    'find_base_target','import_package_to_base','ensure_index_table','base_table_list',
    'list_index_records','archive_index_records','run_lark','_record_fields','_record_id'
])

OUT_DIR=Path('/Users/lilixiaoran/工作/转转/品类数据分析/机型品类周数据_20260427_20260705')
SRC_PACKAGE_ROOT=OUT_DIR/'base_import_packages'
PKG_ROOT=OUT_DIR/'base_import_packages_v2_49k'
RESULT_PATH=OUT_DIR/'base_import_results_v2_49k.json'
RUN_ID='hist_20260427_20260705_v2_49k'
RULE_VERSION='base_partition_v1'
ROW_LIMIT=int(os.environ.get('BASE_PARTITION_ROW_LIMIT','49000'))
TARGETS=[
    ('model','summary','2026-04'),('model','daily_avg','2026-04'),
    ('model','summary','2026-05'),('model','daily_avg','2026-05'),
    ('model','summary','2026-06'),('model','daily_avg','2026-06'),
]
KIND_LABEL={'summary':'汇总','daily_avg':'日均'}
ALIAS={
    '6725f1':'日期机型',
    '7rBBpo':'估价成色',
    '053Pci':'履约',
    'VsIzPj':'估价成色履约',
    'B0ZJKk':'质检成交',
}
PARTITION_INDEX_TABLE_NAME='分片索引'
PARTITION_INDEX_FIELDS=[
    {'name':'记录键','type':'text'},
    {'name':'数据月份','type':'text'},
    {'name':'统计周','type':'text'},
    {'name':'run_id','type':'text'},
    {'name':'状态','type':'text'},
    {'name':'active','type':'checkbox'},
    {'name':'表类型','type':'text'},
    {'name':'source_sheet_id','type':'text'},
    {'name':'逻辑维度','type':'text'},
    {'name':'分片规则版本','type':'text'},
    {'name':'分片字段','type':'text'},
    {'name':'分片序号','type':'text'},
    {'name':'分片总数','type':'number','style': {'type':'plain','precision':0,'thousands_separator':True}},
    {'name':'品类范围','type':'text'},
    {'name':'品类数','type':'number','style': {'type':'plain','precision':0,'thousands_separator':True}},
    {'name':'品类列表摘要','type':'text'},
    {'name':'Base表名','type':'text'},
    {'name':'Base表ID','type':'text'},
    {'name':'行数','type':'number','style': {'type':'plain','precision':0,'thousands_separator':True}},
    {'name':'列数','type':'number','style': {'type':'plain','precision':0,'thousands_separator':True}},
    {'name':'导入时间','type':'datetime','style': {'format':'yyyy-MM-dd HH:mm'}},
    {'name':'备注','type':'text'},
]


def load_results() -> list[dict[str, Any]]:
    if RESULT_PATH.exists():
        return json.loads(RESULT_PATH.read_text(encoding='utf-8'))
    return []


def save_results(results: list[dict[str, Any]]) -> None:
    RESULT_PATH.write_text(json.dumps(results, ensure_ascii=False, indent=2, default=str), encoding='utf-8')


def upsert_result(row: dict[str, Any]) -> None:
    results=load_results()
    key=row.get('key')
    for i, r in enumerate(results):
        if r.get('key') == key:
            results[i]=row
            save_results(results)
            return
    results.append(row)
    save_results(results)


def parse_week(value: Any) -> str:
    if isinstance(value, datetime):
        d=value.date()
    elif isinstance(value, date):
        d=value
    else:
        s=str(value).strip()
        for fmt in ('%Y-%m-%d','%Y/%m/%d','%Y%m%d'):
            try:
                if fmt == '%Y%m%d':
                    d=datetime.strptime(s[:8], fmt).date()
                else:
                    d=datetime.strptime(s[:10], fmt).date()
                break
            except Exception:
                continue
        else:
            raise ValueError(f'cannot parse date {value!r}')
    y,w,_=d.isocalendar()
    return f'{y}-W{w:02d}'


def week_short(week: str) -> str:
    return week.split('-')[1]


def table_name(month: str, week: str, kind: str, source_sheet_id: str, suffix: str | None=None) -> str:
    mm=month[-2:]
    base=f"{mm}_{week_short(week)}_{KIND_LABEL[kind]}_{ALIAS[source_sheet_id]}"
    if suffix:
        base=f'{base}_{suffix}'
    if len(base) > 31:
        raise ValueError(f'table name too long {base} len={len(base)}')
    return base


def normalize_cat(v: Any) -> str:
    if v is None:
        return ''
    return str(v).strip()


def category_digest(categories: list[str]) -> str:
    payload='\n'.join(categories)
    return hashlib.sha1(payload.encode('utf-8')).hexdigest()[:12]


def build_partition_specs(table: dict[str, Any], header: list[Any], counts_week: Counter, counts_week_cat: Counter) -> tuple[dict[tuple[str,str], str], list[dict[str, Any]]]:
    source_sheet_id=table['source_sheet_id']
    month=table['month']
    kind=table['kind']
    # category field: first 品类名称 column. All model logical tables have it.
    specs=[]
    assignment={}
    for wk in sorted(counts_week):
        wk_total=counts_week[wk]
        if wk_total <= ROW_LIMIT:
            name=table_name(month, wk, kind, source_sheet_id)
            spec={
                'month': month,
                'week': wk,
                'kind': kind,
                'source_sheet_id': source_sheet_id,
                'daily_sheet_id': table.get('daily_sheet_id'),
                'business_name': table.get('business_name') or ALIAS[source_sheet_id],
                'base_table_name': name,
                'rows': wk_total,
                'cols': len(header),
                'columns': [str(x) for x in header],
                'partition_rule_version': RULE_VERSION,
                'row_limit': ROW_LIMIT,
                'partition_key': wk,
                'partition_field': '统计周',
                'partition_index': '',
                'partition_total': 1,
                'category_min': '',
                'category_max': '',
                'category_count': 0,
                'category_digest': '',
                'category_list': [],
            }
            specs.append(spec)
            assignment[(wk,'*')]=name
            continue

        cats=sorted([cat for (w,cat), c in counts_week_cat.items() if w == wk], key=lambda x: x)
        buckets=[]
        cur=[]; cur_rows=0
        for cat in cats:
            c=counts_week_cat[(wk,cat)]
            if c > ROW_LIMIT:
                raise RuntimeError(f'single category over row limit: {table["base_table_name"]} {wk} {cat} rows={c}')
            if cur and cur_rows + c > ROW_LIMIT:
                buckets.append((cur, cur_rows))
                cur=[]; cur_rows=0
            cur.append(cat); cur_rows += c
        if cur:
            buckets.append((cur, cur_rows))
        for idx, (bucket_cats, bucket_rows) in enumerate(buckets, start=1):
            suffix=f'C{idx:02d}'
            name=table_name(month, wk, kind, source_sheet_id, suffix)
            for cat in bucket_cats:
                assignment[(wk,cat)] = name
            spec={
                'month': month,
                'week': wk,
                'kind': kind,
                'source_sheet_id': source_sheet_id,
                'daily_sheet_id': table.get('daily_sheet_id'),
                'business_name': table.get('business_name') or ALIAS[source_sheet_id],
                'base_table_name': name,
                'rows': bucket_rows,
                'cols': len(header),
                'columns': [str(x) for x in header],
                'partition_rule_version': RULE_VERSION,
                'row_limit': ROW_LIMIT,
                'partition_key': f'{wk}|{suffix}',
                'partition_field': '品类名称',
                'partition_index': suffix,
                'partition_total': len(buckets),
                'category_min': bucket_cats[0] if bucket_cats else '',
                'category_max': bucket_cats[-1] if bucket_cats else '',
                'category_count': len(bucket_cats),
                'category_digest': category_digest(bucket_cats),
                'category_list': bucket_cats,
            }
            specs.append(spec)
    return assignment, specs


def package_paths(family: str, kind: str, month: str) -> tuple[Path, Path, Path]:
    src_manifest=SRC_PACKAGE_ROOT/f'{family}_{kind}_{month}_hist_20260427_20260705_v1'/'manifest.json'
    pkg=PKG_ROOT/f'{family}_{kind}_{month}_{RUN_ID}'
    manifest=pkg/'manifest.json'
    state=pkg/'state.json'
    return src_manifest,pkg,state


def build_package(family: str, kind: str, month: str) -> dict[str, Any]:
    src_manifest_path,pkg,state_path=package_paths(family,kind,month)
    pkg.mkdir(parents=True, exist_ok=True)
    manifest_path=pkg/'manifest.json'
    if manifest_path.exists():
        manifest=json.loads(manifest_path.read_text(encoding='utf-8'))
        if manifest.get('partition_rule_version') == RULE_VERSION and manifest.get('row_limit') == ROW_LIMIT:
            print('[package-skip]', family, kind, month, manifest_path, flush=True)
            return manifest

    src_manifest=json.loads(src_manifest_path.read_text(encoding='utf-8'))
    source_xlsx=Path(src_manifest['xlsx_path'])
    print('[package-build]', family, kind, month, source_xlsx, flush=True)
    wb=load_workbook(source_xlsx, read_only=True, data_only=True)
    all_specs=[]
    try:
        for table in src_manifest['tables']:
            src_sheet=table['base_table_name']
            ws=wb[src_sheet]
            it=ws.iter_rows(values_only=True)
            header=list(next(it))
            date_idx=header.index('日期') if '日期' in header else 0
            cat_idx=header.index('品类名称') if '品类名称' in header else None
            if cat_idx is None:
                raise RuntimeError(f'品类名称 column missing in {src_sheet}')
            counts_week=Counter(); counts_week_cat=Counter()
            for row in it:
                wk=parse_week(row[date_idx])
                cat=normalize_cat(row[cat_idx])
                counts_week[wk]+=1
                counts_week_cat[(wk,cat)]+=1
            # Re-open worksheet iterator for second pass.
            assignment, specs=build_partition_specs(table, header, counts_week, counts_week_cat)
            print('[partition-plan]', src_sheet, 'source_rows', table['rows'], 'parts', len(specs), 'max_rows', max(s['rows'] for s in specs), flush=True)
            # Create workbooks for partitions for this logical sheet.
            handles={}
            counts=Counter()
            for spec in specs:
                name=spec['base_table_name']
                part_dir=pkg/name
                part_dir.mkdir(parents=True, exist_ok=True)
                xlsx=part_dir/f'{name}.xlsx'
                spec['xlsx_path']=str(xlsx)
                if xlsx.exists():
                    print('[partition-file-skip]', name, xlsx, flush=True)
                    continue
                out_wb=Workbook(write_only=True)
                out_ws=out_wb.create_sheet(title=name)
                out_ws.append(header)
                handles[name]=(out_wb,out_ws,xlsx)
            ws=wb[src_sheet]
            it=ws.iter_rows(values_only=True)
            next(it)
            for row in it:
                wk=parse_week(row[date_idx])
                cat=normalize_cat(row[cat_idx])
                name=assignment.get((wk,cat)) or assignment.get((wk,'*'))
                if not name:
                    raise RuntimeError(f'no assignment for {src_sheet} {wk} {cat}')
                if name in handles:
                    handles[name][1].append(row)
                counts[name]+=1
            for spec in specs:
                name=spec['base_table_name']
                if counts[name] != spec['rows']:
                    raise RuntimeError(f'partition row mismatch {name}: wrote {counts[name]} expected {spec["rows"]}')
            for name,(out_wb,out_ws,xlsx) in handles.items():
                out_wb.save(xlsx)
                print('[partition-file-done]', name, 'rows', counts[name], 'MB', round(xlsx.stat().st_size/1024/1024,2), flush=True)
            all_specs.extend(specs)
    finally:
        wb.close()

    total_rows=sum(int(s['rows']) for s in all_specs)
    if total_rows != int(src_manifest['total_rows']):
        raise RuntimeError(f'total rows mismatch {total_rows} vs source {src_manifest["total_rows"]}')
    over=[s for s in all_specs if int(s['rows']) > ROW_LIMIT]
    if over:
        raise RuntimeError(f'partitions over limit: {[(s["base_table_name"],s["rows"]) for s in over[:5]]}')
    manifest={
        'schema_version': 2,
        'mode': 'base_partition_v1_import_package',
        'family': family,
        'kind': kind,
        'month': month,
        'week': src_manifest.get('week'),
        'run_id': RUN_ID,
        'source_run_id': src_manifest.get('run_id'),
        'created_at': datetime.now().isoformat(timespec='seconds'),
        'partition_rule_version': RULE_VERSION,
        'row_limit': ROW_LIMIT,
        'source_manifest_path': str(src_manifest_path),
        'manifest_path': str(manifest_path),
        'table_count': len(all_specs),
        'logical_table_count': len(src_manifest['tables']),
        'total_rows': total_rows,
        'tables': all_specs,
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
    print('[package-done]', manifest_path, 'tables', len(all_specs), 'rows', total_rows, flush=True)
    return manifest


def load_state(pkg: Path) -> dict[str, Any]:
    p=pkg/'state.json'
    if p.exists():
        return json.loads(p.read_text(encoding='utf-8'))
    return {'imports':{}}


def save_state(pkg: Path, state: dict[str, Any]) -> None:
    (pkg/'state.json').write_text(json.dumps(state, ensure_ascii=False, indent=2, default=str), encoding='utf-8')


def poll_ticket(ticket: str):
    last=None
    for attempt in range(180):
        p=subprocess.run(['lark-cli','drive','+task_result','--scenario','import','--ticket',ticket,'--as','user','--format','json'], text=True, capture_output=True, timeout=120)
        raw=p.stdout or p.stderr
        try: obj=json.loads(raw[raw.index('{'):])
        except Exception: obj={'ok':False,'raw':raw}
        last=obj
        print('[task_result]', ticket, attempt, json.dumps(obj, ensure_ascii=False)[:700], flush=True)
        if obj.get('ok'):
            data=obj.get('data',{})
            label=str(data.get('job_status_label') or data.get('status') or '').lower()
            if data.get('ready') is True or label in {'success','succeeded','done','completed'} or data.get('job_status') in {0,3}:
                return obj
            if data.get('failed') is True or label in {'failed','fail','error'} or data.get('job_status') in {4,5}:
                raise RuntimeError(f'import task failed ticket={ticket}: {obj}')
        time.sleep(20)
    raise TimeoutError(f'import task not ready ticket={ticket} last={last}')


def import_partitions(manifest: dict[str, Any], target: Any) -> dict[str, Any]:
    pkg=Path(manifest['manifest_path']).parent
    state=load_state(pkg)
    state.setdefault('imports',{})
    for spec in manifest['tables']:
        name=spec['base_table_name']
        if state['imports'].get(name,{}).get('status') == 'ready':
            print('[import-skip-ready]', manifest['family'], manifest['kind'], manifest['month'], name, flush=True)
            continue
        xlsx=Path(spec['xlsx_path'])
        print('[import-start]', manifest['family'], manifest['kind'], manifest['month'], name, 'rows', spec['rows'], 'MB', round(xlsx.stat().st_size/1024/1024,2), flush=True)
        data=base_migration.import_package_to_base(xlsx, target.base_token, as_identity='user')
        entry={'status':'submitted','data':data,'xlsx':str(xlsx),'rows':spec['rows']}
        state['imports'][name]=entry; save_state(pkg,state)
        if data.get('ready') is False or data.get('timed_out') is True:
            ticket=data.get('ticket')
            if not ticket:
                raise RuntimeError(f'import pending without ticket {data}')
            entry['task_result']=poll_ticket(ticket)
        entry['status']='ready'
        save_state(pkg,state)
        print('[import-ready]', name, flush=True)
    return state


def table_map(base_token: str) -> dict[str,str]:
    tables=base_migration.base_table_list(base_token, as_identity='user')
    return {str(t.get('name') or t.get('table_name') or t.get('title') or ''): str(t.get('table_id') or t.get('id') or t.get('tableId') or '') for t in tables}


def parse_source_from_record_key(key: Any) -> str | None:
    parts=str(key or '').split('|')
    return parts[2] if len(parts) >= 3 else None


def archive_active_for_month_kind(base_token: str, table_id: str, month: str, kind: str, source_ids: set[str]) -> int:
    records=base_migration.list_index_records(base_token, table_id, as_identity='user')
    ids=[]
    for rec in records:
        fields=rec.get('fields') if isinstance(rec.get('fields'),dict) else rec
        if fields.get('active') is not True:
            continue
        if str(fields.get('数据月份')) != month or str(fields.get('表类型')) != kind:
            continue
        source=parse_source_from_record_key(fields.get('记录键')) or str(fields.get('source_sheet_id') or '')
        if source in source_ids:
            rid=str(rec.get('record_id') or rec.get('id') or rec.get('recordId') or '')
            if rid:
                ids.append(rid)
    if ids:
        return base_migration.archive_index_records(base_token, table_id, ids, as_identity='user')
    return 0


def ensure_partition_index_table(base_token: str) -> str:
    for t in base_migration.base_table_list(base_token, as_identity='user'):
        name=str(t.get('name') or t.get('table_name') or t.get('title') or '')
        if name == PARTITION_INDEX_TABLE_NAME:
            return str(t.get('table_id') or t.get('id') or t.get('tableId'))
    data=base_migration.run_lark(
        'base','+table-create','--base-token',base_token,'--name',PARTITION_INDEX_TABLE_NAME,
        '--fields',json.dumps(PARTITION_INDEX_FIELDS, ensure_ascii=False), as_identity='user'
    )
    token=None
    for k in ('table_id','id','tableId'):
        if data.get(k): token=data[k]
    if not token:
        # use helper's deep finder if exposed indirectly not needed; refresh list.
        return ensure_partition_index_table(base_token)
    return str(token)


def create_rows(base_token: str, table_id: str, fields: list[str], rows: list[list[Any]]) -> int:
    created=0
    for i in range(0,len(rows),200):
        chunk=rows[i:i+200]
        base_migration.run_lark(
            'base','+record-batch-create','--base-token',base_token,'--table-id',table_id,
            '--json',json.dumps({'fields':fields,'rows':chunk}, ensure_ascii=False), as_identity='user'
        )
        created += len(chunk)
    return created


def publish(manifest: dict[str, Any], target: Any, state: dict[str, Any]) -> dict[str, Any]:
    base_token=target.base_token
    pkg=Path(manifest['manifest_path']).parent
    tmap=table_map(base_token)
    expected={s['base_table_name'] for s in manifest['tables']}
    missing=sorted(expected-set(tmap))
    if missing:
        raise RuntimeError(f'missing imported tables {missing[:5]}')
    index_id=base_migration.ensure_index_table(base_token, as_identity='user')
    part_index_id=ensure_partition_index_table(base_token)
    source_ids={s['source_sheet_id'] for s in manifest['tables']}
    archived_main=archive_active_for_month_kind(base_token, index_id, manifest['month'], manifest['kind'], source_ids)
    archived_part=archive_active_for_month_kind(base_token, part_index_id, manifest['month'], manifest['kind'], source_ids)
    imported_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    main_fields=['记录键','统计周','数据月份','run_id','version','状态','active','表类型','业务表名','Base表名','Base表ID','行数','列数','校验结果','导入时间','备注']
    main_rows=[]
    for s in manifest['tables']:
        category_info=''
        if s.get('partition_field') == '品类名称':
            category_info=f";品类范围={s.get('category_min','')}~{s.get('category_max','')};品类数={s.get('category_count',0)};摘要={s.get('category_digest','')}"
        note=f"drive_import_bitable;rule={RULE_VERSION};limit={ROW_LIMIT};partition={s.get('partition_key','')};field={s.get('partition_field','')}{category_info}"
        key=f"{s['week']}|{s['kind']}|{s['source_sheet_id']}|{s.get('partition_key','')}|{manifest['run_id']}"
        main_rows.append([key,s['week'],manifest['month'],manifest['run_id'],manifest['run_id'],'已发布',True,s['kind'],s['business_name'],s['base_table_name'],tmap[s['base_table_name']],int(s['rows']),int(s['cols']),'manifest_matched',imported_at,note])
    created_main=create_rows(base_token,index_id,main_fields,main_rows)

    part_fields=['记录键','数据月份','统计周','run_id','状态','active','表类型','source_sheet_id','逻辑维度','分片规则版本','分片字段','分片序号','分片总数','品类范围','品类数','品类列表摘要','Base表名','Base表ID','行数','列数','导入时间','备注']
    part_rows=[]
    for s in manifest['tables']:
        key=f"{s['week']}|{s['kind']}|{s['source_sheet_id']}|{s.get('partition_key','')}|{manifest['run_id']}"
        cat_range=f"{s.get('category_min','')}~{s.get('category_max','')}" if s.get('category_min') or s.get('category_max') else ''
        note=f"category_list={','.join(s.get('category_list',[])[:20])}" if s.get('category_list') else ''
        part_rows.append([key,manifest['month'],s['week'],manifest['run_id'],'已发布',True,s['kind'],s['source_sheet_id'],s['business_name'],RULE_VERSION,s.get('partition_field',''),s.get('partition_index',''),int(s.get('partition_total') or 1),cat_range,int(s.get('category_count') or 0),s.get('category_digest',''),s['base_table_name'],tmap[s['base_table_name']],int(s['rows']),int(s['cols']),imported_at,note])
    created_part=create_rows(base_token,part_index_id,part_fields,part_rows)
    result={'status':'published','base_token':base_token,'index_table_id':index_id,'partition_index_table_id':part_index_id,'archived_main':archived_main,'archived_partition_index':archived_part,'created_main':created_main,'created_partition_index':created_part,'base_tables':{name:tmap[name] for name in expected}}
    state['publish']=result; save_state(pkg,state)
    print('[publish-done]', manifest['family'], manifest['kind'], manifest['month'], result, flush=True)
    return result


def run_one(family: str, kind: str, month: str) -> None:
    target=base_migration.find_base_target(month, kind, family=family)
    if not target:
        raise RuntimeError(f'missing target {family} {kind} {month}')
    manifest=build_package(family,kind,month)
    state=import_partitions(manifest,target)
    if state.get('publish',{}).get('status') == 'published':
        print('[publish-skip-ready]', family, kind, month, flush=True)
        pub=state['publish']
    else:
        pub=publish(manifest,target,state)
    upsert_result({'key':[family,kind,month],'status':'published','target_label':target.label,'target_url':target.url,'manifest':manifest['manifest_path'],'result':pub})


def main():
    requested=sys.argv[1:]
    targets=TARGETS
    if requested:
        req=set(tuple(x.split('/')) for x in requested)
        targets=[t for t in TARGETS if t in req]
    for t in targets:
        print('\n=== RUN', t, '===', flush=True)
        run_one(*t)
    print('[all-done]', RESULT_PATH, flush=True)

if __name__ == '__main__':
    main()
